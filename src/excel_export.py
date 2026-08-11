from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def _style_header(ws):
    fill = PatternFill(fill_type="solid", fgColor="EA580C")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    return fill


def _make_table(ws, name):
    if ws.max_row < 2:
        return
    from openpyxl.utils import get_column_letter
    end = get_column_letter(ws.max_column)
    table = Table(displayName=name, ref=f"A1:{end}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def _autosize(ws) -> None:
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter
    for column_index in range(1, ws.max_column + 1):
        max_length = 0
        for row_index in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_index, column=column_index)
            if isinstance(cell, MergedCell):
                continue
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        column_letter = get_column_letter(column_index)
        ws.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 40)


def export_invoices_to_excel(
    invoices: List[Dict[str, Any]],
    output_path: Path,
    line_items: Optional[List[Dict[str, Any]]] = None,
    vendor_summary: Optional[List[Dict[str, Any]]] = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Invoices"
    invoice_headers = ["Source File","Vendor","Invoice Number","Invoice Date","Due Date","Description","Subtotal","Tax","Shipping","Total Due","Validation Status","Validation Errors","Created At"]
    ws.append(invoice_headers)
    keys = ["source_file","vendor","invoice_number","invoice_date","due_date","description","subtotal","tax","shipping","total_due","validation_status","validation_errors","created_at"]
    for inv in invoices:
        ws.append([inv.get(k) for k in keys])
    _style_header(ws)
    for col in ["G","H","I","J"]:
        for cell in ws[col][1:]:
            cell.number_format = "$#,##0.00"
    _make_table(ws, "InvoicesTable")
    _autosize(ws)

    li = wb.create_sheet("Line Items")
    li.append(["Vendor","Invoice Number","Description","Quantity","Unit Price","Line Total"])
    if line_items:
        for item in line_items:
            li.append([item.get("vendor"), item.get("invoice_number"), item.get("description"), item.get("quantity"), item.get("unit_price"), item.get("line_total")])
    _style_header(li)
    for col in ["E","F"]:
        for cell in li[col][1:]:
            cell.number_format = "$#,##0.00"
    _make_table(li, "LineItemsTable")
    _autosize(li)

    vs = wb.create_sheet("Vendor Summary")
    vs.append(["Vendor","Invoices","Total Spend","Average Invoice","Largest Invoice"])
    if vendor_summary:
        for row in vendor_summary:
            vs.append([row.get("vendor"), row.get("invoice_count"), row.get("total_spend"), row.get("average_invoice"), row.get("largest_invoice")])
    _style_header(vs)
    for col in ["C","D","E"]:
        for cell in vs[col][1:]:
            cell.number_format = "$#,##0.00"
    _make_table(vs, "VendorSummaryTable")
    _autosize(vs)

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Invoice AI — Ledger Export"
    summary["A1"].font = Font(size=18, bold=True)
    summary.merge_cells("A1:B1")
    total = sum(float(i.get("total_due") or 0) for i in invoices)
    valid = sum(1 for i in invoices if i.get("validation_status") == "Valid")
    rows = [("Invoices", len(invoices)), ("Total Spend", total), ("Validated", valid), ("Needs Review", len(invoices) - valid)]
    r = 3
    for label, value in rows:
        summary[f"A{r}"] = label
        summary[f"B{r}"] = value
        if isinstance(value, (int, float)) and "Spend" in label:
            summary[f"B{r}"].number_format = "$#,##0.00"
        r += 1
    _autosize(summary)

    wb.save(output_path)